# Import necessary libraries
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Color, colors, fills
import os
from os import walk
import win32com.client
import time
from win32com.client import constants
from datetime import datetime
import warnings
from datetime import date

# Suppresses all warnings
warnings.filterwarnings('ignore')

organisms = ["Campylobacter","Escherichia", "Listeria" ,"Salmonella", "Vibrio"] 
path_to_downloads = "//kdhe/dfs/LabShared/Molecular Genomics Unit/Testing/PulseNet/Downloaded data/"
path_to_epi_report = "//kdhe/dfs/EPI/LAB_OSE/WGS/"      
json_path = "cluster_tracker.json"
path_to_results = "//kdhe/dfs/epi/lab_ose/wgs/script_results/"    


cutoff = 10

# Function to extract the date column from the given patient DOB and Sample ID
def get_attr(row, attr, df):
    if attr == 'PatientDOB':
        date = df.loc[row['Sample ID'], attr]
        print(date)
        return date.date()
    else:
        return df.loc[row['Sample ID'], attr]

        
# Function to extract numeric part from column names
def extract_numeric_part(col_name):
    if 'KS___' in col_name:
        return col_name.split('KS___')[-1]  # Get the part after 'KS___'
    return col_name

# This function shades/colored the sheets values based on the lowest and highest value.
def shade_workbooks(path_lst):
    print("\nColoring workbooks...")
    xl = win32com.client.gencache.EnsureDispatch("Excel.Application")
    xl.ScreenUpdating = False

    for path in path_lst:
        try:
            workbook = xl.Workbooks.Open(path)

            for sheet in workbook.Sheets:
                if sheet.Name not in ['Summary', 'Summary_Demographics']:
                    xl.Worksheets(sheet.Name).Activate()
                    ws = xl.ActiveSheet
                    
                    # Select the specified range and add color scale
                    ws.Range("B2:GS201").Select()
                    
                    # Add color scale
                    color_scale = xl.Selection.FormatConditions.AddColorScale(ColorScaleType=2)
                    color_scale.SetFirstPriority()
                    color_scale.ColorScaleCriteria(1).Type = constants.xlConditionValueLowestValue
                    color_scale.ColorScaleCriteria(1).FormatColor.Color = 8109667
                    color_scale.ColorScaleCriteria(1).FormatColor.TintAndShade = 0

                    color_scale.ColorScaleCriteria(2).Type = constants.xlConditionValueHighestValue
                    color_scale.ColorScaleCriteria(2).FormatColor.Color = 10285055
                    color_scale.ColorScaleCriteria(2).FormatColor.TintAndShade = 0
                    
        except Exception as e:
            print(f"Error processing workbook {path}: {e}")
        finally:
            workbook.Close(SaveChanges=True)
    xl.Application.Quit()


# This function creates the outbreak dataframe based on the outbreak column ig??
def create_outbreak_df(df_of_outbreaks, columns_summary,org):
    outbreak_summary=[]
    
    for index,row in df_of_outbreaks.iterrows():
        temp_summary = pd.DataFrame(columns=columns_summary)
        if org == 'Escherichia':
            # print([index,row[21],row[22],row[96],row[27],row[23],row[31],row[26],row[13],row[6] ])

            temp_summary.loc[0] = [index,row[21],row[22],row[96],row[27],row[23],row[31],row[26],row[13],row[6] ]

            temp_summary.rename(columns={'Sample ID':row[6]}, inplace=True)
            temp_summary= temp_summary.set_index(row[6])
         
            temp_summary.style.set_caption(row[6])
        elif org == "Listeria" : 
            
            temp_summary.loc[0] = [index,row[18],row[19],'',row[20],row[25],row[28],row[23],row[14],row[12] ]

       
            temp_summary.rename(columns={'Sample ID':row[12]}, inplace=True)
            temp_summary= temp_summary.set_index(row[12])
         
            temp_summary.style.set_caption(row[12])


        else:

            # print([index,row[9],row[10],row[13],row[11],row[14],row[18],row[17],row[22],row[4]])
            temp_summary.loc[0] = [index,row[9],row[10],row[13],row[11],row[14],row[18],row[17],row[22],row[4] ]
       
            temp_summary.rename(columns={'Sample ID':row[4]}, inplace=True)
            temp_summary= temp_summary.set_index(row[4])
         
            temp_summary.style.set_caption(row[4])

        outbreak_summary.append(temp_summary)
        # print(outbreak_summary)

    return outbreak_summary

# This function checks if the hsn is present in the previous reports or not and if not then append the new hsn to a list.
def check_if_in_previous_report(sample_hsns,path_to_res,current_organism,curr_run_date):

    dates= datetime.strptime("010100","%m%d%y")
    # Find the max date
    for file in next(walk(path_to_res), (None, None, []))[2] :
        if file[0].isnumeric():
            temp_date = datetime.strptime(file.split(" ")[0],"%m%d%y")
            if temp_date != datetime.strptime(curr_run_date,"%m%d%y") and temp_date >= dates :
                dates = temp_date
    # First find closesest date to current date from list of folders below
    dates= dates.strftime("%m%d%y")
    # Then read only the Summary sheet of that execl file
    print("\n\n\n")
    print("using this as previous date for Epi Report"+ dates)
    try:
        old_results = pd.read_excel(path_to_res+dates+" "+current_organism+" clusters.xlsx",sheet_name='Summary')
        column= old_results.columns.values.tolist()
        column_values= old_results[column[0]].iloc[1:].dropna().astype(str).tolist()
        column_values = [str(i) for i in column_values]
        # Then go line by line to to turn that first column into the list

        new_hsns=[]
        # Then loop through hsns to see if they exsist
        for hsn in sample_hsns:
            if hsn not in column_values:
                print("this sample is NEW "+hsn)
                new_hsns.append(hsn)
        return new_hsns
    except:
        print("no prev culsters for "+ current_organism)
        return []      

# This function highlights the hsn which exsist i.e. the hsn is new and is not present in any of prev clusters.
# If they exsist then find a way to highlight them!
def highlight_newrows(path_to_outputfile, new_hsns):
    wb = openpyxl.load_workbook(path_to_outputfile)
    summary_worksheet = wb['Summary']
    for row in summary_worksheet.iter_rows(max_col=1):
        for cell in row:
            if str(cell.value) in new_hsns:  # Ensure HSNs are compared as strings
                cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow fill
    wb.save(path_to_outputfile)


# This function finds and return data related to specific serotypes of Salmonella from a given DataFrame. 
# It filters the DataFrame to find entries corresponding to certain serotypes and organizes these filtered entries in a dictionary. 
def find_serotype(salmonella_df):
    c_name = "Serotype_wgs"
    # typhi, paratyphi A, paratyphi B tartrate negative, and paratyphi C
    cdc = ["Typhi","Paratyphi A","Paratyphi B","Paratyphi C"] 
    #S typhi, Paratyphi A, B (tartrate negative), and C
    serotype_dfs={}
    #temp_df.query("Outbreak.notnull()")

    for sero_type in cdc:
        temp_df= salmonella_df.query(c_name+" == @sero_type")
        if not(temp_df.empty):
            serotype_dfs[sero_type]= temp_df  # The dictionary's keys are the serotype names, and the values are DataFrames with rows matching each serotype.
            #print("for "+sero_type)
            #print("-"*100)
            #print(serotype_dfs[sero_type])
            #print("-"*100)
    return serotype_dfs


# Salmonella Dataset
def format_df_sal(p,first_df,r_date):
    rename_col_lst= {
    "Key":"HSN",
    "WGS_id":"WGS_ID",
    "FirstName": "Patient_First_Name",
    "LastName": "Patient_Last_Name",
    "PatientDOB": "Patient_DOB",
    "PatientSex": "Patient_Gender",
    "SourceCounty":"County",
    "SourceState":"State",
    "LabID": "Lab",
    "Serotype_wgs": "Serotype_wgs",  
    "NCBI_ACCESSION":"NCBI_ACCESSION",
    "Allele_Code":"Allele_Code",
    "Outbreak":"Outbreak",
    "REP_code":"REP_code",  
    "PulseNet_UploadDate":"PulseNet_UploadDate",
    "IsolatDate": "Collection_Date",
    "SourceSite":"Specimen_Source"}

    csv_headers= [
    "HSN",
    "WGS_ID",
    "Patient_Last_Name",
    "Patient_First_Name",
    "Patient_DOB",
    "Patient_Gender",
    "County",
    "State",
    "Lab",
    "Serotype_wgs",
    "NCBI_ACCESSION",
    "Allele_Code",
    "Outbreak",
    "REP_code",
    "PulseNet_UploadDate",
    "Collection_Date",
    "Specimen_Source"]

    #sort through the Key 
    #if Key.isnumeric
    main_df = first_df[pd.to_numeric(first_df['Key'], errors='coerce').notnull()]

    # drop others
    main_df = main_df[[*rename_col_lst]]

    #rename
    main_df = main_df.rename(columns = rename_col_lst)
    
    #sort order 
    main_df = main_df[csv_headers]

    #remove non KS state from DF
    main_df = main_df[main_df["Lab"] == "KS"]

    curr_hsn = main_df["HSN"].values.tolist()
    curr_hsn = [str(i) for i in curr_hsn]

    new_hsn = keep_only_new_records(p+"Epi_Track_Output/Salmonella/",curr_hsn,r_date)
    
    new_hsn = [int(i) for i in new_hsn]
    main_df = main_df[main_df["HSN"].isin(new_hsn)]
    
    # write to file
    main_df.to_csv(p+"Epi_Track_Output/Salmonella/"+r_date+"_epiTrackOutput_Salmonella.csv",index=False)

# This function keeps only new HSNs after we check for the HSNs from the recent previous report or the old_hsn.csv file.
def keep_only_new_records(path,current_hsn,curr_run_date):
    
    # Open old_hsn.csv

    try:
        old_upload = pd.read_csv(path+"old_hsn.csv",header=0)
        
        column_values= old_upload[old_upload.columns.values.tolist()[0]].values.tolist()
        # Then go line by line to to turn that first column into the list
        column_values = [str(i) for i in column_values]
        new_hsns=[]
        # Then loop through hsns to see if they exist 
        for hsn in current_hsn:
            if hsn not in column_values:
                print("this sample is NEW "+hsn)
                new_hsns.append(hsn)

                old_upload.loc[len(old_upload.index)] = [str(hsn)]  # This line is effectively appending a new row to the DataFrame old_upload with the value of hsn converted to a string.

        old_upload.to_csv(path+"old_hsn.csv",index=False)
        # Return the HSNs which will be pushed to the excel file.
        return new_hsns

    except: 
        return []


# Escherichia Dataset
def format_df_ecoli(p,first_df,r_date):
    rename_col_lst= {
    "Key":"HSN",
    "WGS_id":"WGS_ID",
    "FirstName": "Patient_First_Name",
    "LastName": "Patient_Last_Name",
    "PatientDOB": "Patient_DOB",
    "PatientSex": "Patient_Gender",
    "SourceCounty":"County",
    "SourceState":"State",
    "LabID": "Lab",
    "Serotype_wgs": "Serotype_wgs",  
    "NCBI_ACCESSION":"NCBI_ACCESSION",
    "Allele_Code":"Allele_Code",
    "Outbreak":"Outbreak",
    "REP_code":"REP_code",  
    "PulseNet_UploadDate":"PulseNet_UploadDate",
    "IsolatDate": "Collection_Date",
    "SourceSite":"Specimen_Source",
    "Toxin_wgs":"Toxin_wgs",
    "Escherichia_group":"Escherichia_group"}

    csv_headers= [
    "HSN",
    "WGS_ID",
    "Patient_Last_Name",
    "Patient_First_Name",
    "Patient_DOB",
    "Patient_Gender",
    "County",
    "State",
    "Lab",
    "Serotype_wgs",
    "NCBI_ACCESSION",
    "Allele_Code",
    "Outbreak",
    "REP_code",
    "PulseNet_UploadDate",
    "Collection_Date",
    "Specimen_Source",
    "Toxin_wgs",
    "Escherichia_group"]

    #sort through the Key 
    #if Key.isnumeric
    main_df = first_df[pd.to_numeric(first_df['Key'], errors='coerce').notnull()]

    # drop others
    main_df = main_df[[*rename_col_lst]]

    #rename
    main_df = main_df.rename(columns = rename_col_lst)
    
    #sort order 
    main_df = main_df[csv_headers]

    #remove non KS state from DF
    main_df = main_df[main_df["Lab"] == "KS"]

    curr_hsn = main_df["HSN"].values.tolist()
    curr_hsn = [str(i) for i in curr_hsn]

    new_hsn = keep_only_new_records_ecoli(p+"Epi_Track_Output/Escherichia/",curr_hsn,r_date)
    
    new_hsn = [int(i) for i in new_hsn]
    main_df = main_df[main_df["HSN"].isin(new_hsn)]
    
    # write to file
    main_df.to_csv(p+"Epi_Track_Output/Escherichia/"+r_date+"_epiTrackOutput_Escherichia.csv",index=False)


# This function keeps only new HSNs after we check for the HSNs from the recent previous report or the old_hsn_ecoli.csv file.
def keep_only_new_records_ecoli(path,current_hsn,curr_run_date):
    
    # Open old_hsn.csv

    try:
        old_upload = pd.read_csv(path+"old_hsn_ecoli.csv",header=0)
        
        column_values= old_upload[old_upload.columns.values.tolist()[0]].values.tolist()
        # Then go line by line to to turn that first column into the list
        column_values = [str(i) for i in column_values]
        new_hsns=[]
        # Then loop through hsns to see if they exist 
        for hsn in current_hsn:
            if hsn not in column_values:
                print("this sample is NEW "+hsn)
                new_hsns.append(hsn)

                old_upload.loc[len(old_upload.index)] = [str(hsn)]  # This line is effectively appending a new row to the DataFrame old_upload with the value of hsn converted to a string.

        old_upload.to_csv(path+"old_hsn_ecoli.csv",index=False)
        # Return the HSNs which will be pushed to the excel file.
        return new_hsns

    except: 
        return []

if __name__ == "__main__":

    # Read in the cluster tracker data, we need to find out which cluster the samples belong to.

    # Get the date from the user
    run_date = input("\nPlease enter the date of the download you made in mmddyy format\n--> ")
    demo_result_name = run_date + " Epi report past 90.xlsx"

    # Open up the demographics matrix:
    demo_path = path_to_epi_report+demo_result_name
    matrix_path_base = "/".join(demo_path.split("/")[:-1])
    organism_demo_df_dict = {}
    all_samples_found={}

    # Create a dictionary of dataframes for each organism
    for organism in organisms:
        try:
            df = pd.read_excel(demo_path, sheet_name=organism)
            df['Key'] = df['Key'].astype(str)
            df = df.set_index('Key')
            # Should check if serotype here after reading in files
            if organism == "Salmonella":
                serotype_df=find_serotype(df)

            organism_demo_df_dict[organism] = [df]
        except:
            pass
        
    # Read the matrices into dataframe for analysis
    for organism in organisms:
        path_matrix = matrix_path_base + "/" + run_date + " matrix " + organism + ".xlsx"
        try:
            df = pd.read_excel(path_matrix)
            organism_demo_df_dict[organism].append(df)
        except:
            print("failed opening matrix "+organism)

    # Capture the matrices, generating a new dataframe for each one
    epi_matrices = {}
    used_aa_codes={}
    for organism in organism_demo_df_dict.keys():
        if organism == 'Salmonella':
            cutoff = 5
        else:
            cutoff = 10
        print("Collecting matrices for " + organism)
        epi_matrices[organism] = {}

        try:
            matrix_df = pd.DataFrame(organism_demo_df_dict[organism][1])
            
            # Make HSN index
            matrix_df = matrix_df.rename(columns={'samples': 'Key'})
            matrix_df['Key'] = matrix_df['Key'].astype(str)
            matrix_df = matrix_df.set_index('Key')

            # Clean column names and index names with specific prefix
            matrix_df.columns = [extract_numeric_part(col) for col in matrix_df.columns]
            matrix_df.index = [extract_numeric_part(idx) for idx in matrix_df.columns]
            
            demo_df = pd.DataFrame(organism_demo_df_dict[organism][0])
    
            index = 0
            end=False
            while index < len(matrix_df.index):
                # We wish to examine the rows below the diagonal line path we are taking
                center_val = matrix_df.iloc[index, index]
                try:
                    lower_val = matrix_df.iloc[index, index+1]
                except IndexError:
                    break
                if lower_val <= cutoff:
                    
                    # This organism is in a group with the one to the right of it
                    # We need to find the allele code for these, and create a dict if it does not already exist
                    lower_val = matrix_df.iloc[index, index+2]
                    offset = 2
                    # Move downwards, trying to capture index of the end of the cluster
                    while lower_val <=cutoff:
                        offset+=1
                        try:
                            lower_val = matrix_df.iloc[index, index+offset]
                           
                        except IndexError:
                            end=True
                            break
                    if end:
                        break
                    curr_idx = list(range(index, index+offset))
                    # Pull the smaller matrix out of the larger one
                    current_matrix = matrix_df.iloc[curr_idx, curr_idx]
                    # Don't examine the same samples again
                    index += offset
                    # Only consider clusters >= size 3 
                    if not len(current_matrix.index) < 2:
                        # Determine the amino acid code shared by all elements
                        # Get list of elements
                        keys = [str(x) for x in current_matrix.columns]
                        demo_keys = list(demo_df.index)
                        aa_codes = [demo_df.loc[x, 'Allele_Code'] for x in keys]
                        aa_code = str(aa_codes[0])
                        for key in aa_codes[1:]:
                            ctr=0
                            
                            key = str(key)
                            while ctr < len(str(aa_code)) and ctr < len(str(key)):

                                if not str(aa_code[ctr]) == str(key[ctr]):
                                    aa_code = aa_code[:ctr]
                                    break
                                ctr += 1
                        
                        if aa_code.count('.') < 6:
                            print("AA code less then 6")
                            print(aa_code)

                            aa_code += "x"
                        current_matrix.style.set_caption(aa_code)
                        # Add the matrix to the dictionary
                        #print(current_matrix)
                        
                        if aa_code in used_aa_codes.keys():
                            used_aa_codes[aa_code]+=1
                            aa_code+="_"+str(used_aa_codes[aa_code])

                            epi_matrices[organism][aa_code] = current_matrix
                        else:
                            used_aa_codes[aa_code]=1
                            epi_matrices[organism][aa_code] = current_matrix
                        #print(aa_code)    
                        #print(epi_matrices[organism][aa_code])
                else:
                    index+=1
        except IndexError:
            pass

    # Need to loop through and check if something has an outbreak code
    # organism_demo_df_dict[organism][0] check this df
    outbreaks ={}
    for organism in organisms:
        # Check if the organism's demo Dataframe exists 
        if organism in organism_demo_df_dict:
            temp_df = organism_demo_df_dict[organism][0]

            # Using .query() method
            if 'Outbreak' in temp_df.columns:
                outbreaks[organism] = temp_df.query("Outbreak.notnull()")
            else:
                print(f"'Outbreak' column not found for {organism}.")
        else:
            print(f"No demo data found for {organism}. Skipping...")
            continue  # Skip to the next organism if not found

    summaries = {}

    col_order = ['Sample ID', 'LastName', 'FirstName', 'SourceCounty', 'PATIENTAGEYEARS', 'PatientSex', 'SourceSite','PulseNet_UploadDate','Outbreak']
    # format the dataframes  
    for organism in epi_matrices.keys():
        if organism not in organisms:
            print(f"Skipping {organism} as it is not in the predefined list.")
            continue  # Skip organisms not in the predefined list
        summaries[organism] = []
        all_samples_found[organism]=[]
        for aa_code in epi_matrices[organism].keys():
            current_matrix = epi_matrices[organism][aa_code]
            # Create summary DataFrame
            summary_matrix = pd.DataFrame(list(current_matrix.index), columns= ["Sample ID"])
            # Adding columns conditionally 
            summary_matrix['LastName'] = summary_matrix.apply(lambda row: get_attr(row, "LastName", organism_demo_df_dict[organism][0]), axis=1)
            summary_matrix['FirstName'] = summary_matrix.apply(lambda row: get_attr(row, "FirstName", organism_demo_df_dict[organism][0]), axis=1)
            summary_matrix['SourceCounty'] = summary_matrix.apply(lambda row: get_attr(row, "SourceCounty", organism_demo_df_dict[organism][0]), axis=1)
            summary_matrix['PATIENTAGEYEARS'] = summary_matrix.apply(lambda row: get_attr(row, "PATIENTAGEYEARS", organism_demo_df_dict[organism][0]), axis=1)
            if organism in ['Salmonella', 'Escherichia']:
                summary_matrix['PatientDOB'] = summary_matrix.apply(lambda row: get_attr(row, 'PatientDOB', organism_demo_df_dict[organism][0]), axis=1)
                if 'PatientDOB' not in col_order:
                    col_order.insert(3, 'PatientDOB')
            else:
                pass
            summary_matrix['PatientSex'] = summary_matrix.apply(lambda row: get_attr(row, "PatientSex", organism_demo_df_dict[organism][0]), axis=1)
            summary_matrix['SourceSite'] = summary_matrix.apply(lambda row: get_attr(row, "SourceSite", organism_demo_df_dict[organism][0]), axis=1)
            summary_matrix['PulseNet_UploadDate'] = summary_matrix.apply(lambda row: get_attr(row, "PulseNet_UploadDate", organism_demo_df_dict[organism][0]), axis=1)
            summary_matrix['Outbreak'] = summary_matrix.apply(lambda row: get_attr(row, "Outbreak", organism_demo_df_dict[organism][0]), axis=1)

            # Recreate summary_matrix based on unique columns
            summary_matrix = summary_matrix[col_order] # Ensure unique columns and reorder
            summary_matrix.rename(columns={'Sample ID':aa_code}, inplace=True)
            summary_matrix = summary_matrix.set_index(aa_code)
            summaries[organism].append(summary_matrix)  # Append unstyled matrix or a styled version
            
            all_samples_found[organism]+= list(current_matrix.index)
            #print("current summary_matrix")
            #print(summary_matrix)
            for sample in list(current_matrix.index):   
                if sample in outbreaks[organism].index.tolist():
                    print(sample +" was found in another cluster removing from outbreaks")
                    # Need to remove this sample from this table
                    outbreaks[organism] = outbreaks[organism].drop(sample)     
        
        # Create outbreak into summary format
        if organism in outbreaks:
            all_samples_found[organism]+= outbreaks[organism].index.tolist()
            # print("outbreaks found for "+organism)
            # print(outbreaks[organism])
            outbreak_summaries = create_outbreak_df(outbreaks[organism],col_order,organism)
            # Add outbreak into summary
            summaries[organism]+= outbreak_summaries
        else:
            print(f"No outbreak data found for {organism}.")


    # Write every result to a separate sheet within the organism's workbook
    workbook_lst = [] # SOMEWHERE HERE CREATE TAB WITH OUTPUT FROM SEROTYPE
    print("\nCreating Workbooks...")
    for organism in epi_matrices:
        workbook_path = path_to_results + run_date + " " + organism + " clusters" + ".xlsx"
        row_offset = 0
        col_offset = 0

        # Create an empty Dataframe to initialize the workbook
        data = pd.DataFrame(None)
        data.to_excel(workbook_path, index=False)
        
        with pd.ExcelWriter(workbook_path, mode='a',engine="openpyxl", if_sheet_exists= 'overlay') as writer:
            for aa_code in epi_matrices[organism].keys():
                # print(aa_code)
                current_matrix = epi_matrices[organism][aa_code]
    
                # Reset the index to convert the index into a column
                current_matrix.reset_index(inplace=True)

                # Rename the first column to 'Key'
                current_matrix.rename(columns={'index': 'Key'}, inplace=True)
                # Set 'Key' as the index before writing to Excel
                current_matrix.set_index('Key', inplace=True)
                # Write DataFrame to Excel
                current_matrix.to_excel(writer, sheet_name=(aa_code.replace(":","")[:31]),index = True)  
            
            # Write summaries for every organism
            for cluster in summaries[organism]:
                cluster.to_excel(writer, sheet_name="Summary", startrow=row_offset, startcol=col_offset)
                row_offset += 2 + len(cluster.index)

            # Write demographic information
            all_info_df = organism_demo_df_dict[organism][0].loc[all_samples_found[organism]]
            all_info_df.to_excel(writer,sheet_name="Summary_Demographics")

            # Write serotype data if applicable
            if organism == "Salmonella":
                for k in [*serotype_df] :
                    serotype_df[k].to_excel(writer,sheet_name="Serotype")
            # Call the check_if_in_previous_report() to get the new hsn which will go to the final output file.
            new_hsn = check_if_in_previous_report(all_samples_found[organism], path_to_results, organism, run_date)

        # Clean up workbook
        workbook = openpyxl.load_workbook(workbook_path)
        if workbook.sheetnames == ['Sheet1']:
            workbook.close()
            os.remove(workbook_path)
        else:
            del workbook['Sheet1']
            workbook.save(workbook_path)
            workbook.close()
            workbook_lst.append(workbook_path) 
        if new_hsn != []:
            highlight_newrows(workbook_path,new_hsn) # Highlight the new hsn row by calling highlight_newrows().
        
    shade_workbooks(workbook_lst) # Calling shade_workbooks() to get the shaded workbook/excel file!

    # Create epi tracks output
    # For Salmonella samples
    salmonella_df = pd.read_excel(demo_path, sheet_name="Salmonella")
    # For Escheria Coli samples.
    Escherichia_df = pd.read_excel(demo_path, sheet_name="Escherichia")
    # Use the format_df function to format each DataFrame
    format_df_sal(path_to_results,salmonella_df,run_date)
    format_df_ecoli(path_to_results,Escherichia_df,run_date)
